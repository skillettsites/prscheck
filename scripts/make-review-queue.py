"""Turn harvester evidence into a human curation queue (.xlsx).

The harvester deliberately never writes scheme records: with search enabled it
reaches 100% recall against the 90 already-curated councils but only ~0/8
precision on councils that merely *discuss* selective licensing (consultations,
rejected proposals, quashed designations all read the same to a regex). So the
output is a worklist, and a person decides.

One row per council with evidence, ordered so the strongest signals come first.
Columns mirror the index-priority sheet: a Done? and Date done to fill in.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path(sys.argv[1] if len(sys.argv) > 1 else "evidence-final.json")
OUT = Path(sys.argv[2] if len(sys.argv) > 2 else "council-review-queue.xlsx")

ORDER = {"likely_has_scheme": 0, "needs_review": 1, "likely_no_scheme": 2, "no_evidence": 3}
CONF = {"high": 0, "medium": 1, "low": 2, "none": 3}

rows = json.loads(SRC.read_text(encoding="utf-8"))
withev = [r for r in rows if r.get("evidence")]
withev.sort(key=lambda r: (ORDER.get(r["verdict"], 9), CONF.get(r["confidence"], 9), r["name"]))

wb = Workbook()
ws = wb.active
ws.title = "Review queue"

headers = [
    "#", "Council", "Verdict", "Confidence", "Why it was flagged (quoted from the page)",
    "Source URL", "Years seen", "Fees seen", "Active scheme? (y/n)", "Type", "Done?", "Date done", "Notes",
]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="1E3A5F")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = head_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

band = {
    "likely_has_scheme": PatternFill("solid", fgColor="FFF2CC"),
    "needs_review": PatternFill("solid", fgColor="F2F2F2"),
    "likely_no_scheme": PatternFill("solid", fgColor="E8F1E8"),
}

for i, r in enumerate(withev, 1):
    ev = r["evidence"][0]
    quote = next((q for e in r["evidence"] for q in e.get("quotes", []) if q), "")
    ws.append([
        i,
        r["name"],
        r["verdict"].replace("_", " "),
        r["confidence"],
        quote[:400],
        ev.get("url", ""),
        ", ".join(ev.get("years", [])[:6]),
        ", ".join(ev.get("fees", [])[:4]),
        "",  # for the reviewer
        "",  # selective / additional / none
        "", "", "",
    ])
    fill = band.get(r["verdict"])
    if fill:
        for c in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=c).fill = fill

widths = [5, 26, 18, 11, 70, 60, 16, 16, 16, 14, 9, 12, 30]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"

# Second sheet: councils where nothing was found, so the gap is visible rather
# than implied. These are candidates for a genuine "no scheme" record only
# after a spot-check, never by default.
ws2 = wb.create_sheet("No evidence found")
ws2.append(["#", "Council", "Discovery used", "Pages fetched", "Checked? ", "Has scheme? (y/n)", "Notes"])
for c in range(1, 8):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = head_fill
none = [r for r in rows if not r.get("evidence")]
none.sort(key=lambda r: r["name"])
for i, r in enumerate(none, 1):
    ws2.append([i, r["name"], str(r.get("discovery")), r.get("pagesFetched", 0), "", "", ""])
for c, w in enumerate([5, 30, 18, 14, 12, 18, 34], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.freeze_panes = "A2"

wb.save(OUT)
print(f"  {OUT}")
print(f"  review queue rows: {len(withev)}   no-evidence rows: {len(none)}")
